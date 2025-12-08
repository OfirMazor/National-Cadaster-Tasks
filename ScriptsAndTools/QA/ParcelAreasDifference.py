"""
Produces an Excel spreadsheet containing the difference between a parcel's stated and computed areas,
and whether that difference is valid or not as defined in the Survey of Israel's 2016 regulations.
https://www.gov.il/he/pages/laws-and-regulations1
"""
# https://joelmccune.com/relative-module-imports-in-an-arcgis-python-toolbox/
import os
import sys
sys.path.append(os.path.abspath(".."))

import arcpy
from arcpy import Exists, AddMessage
from arcpy.da import SearchCursor, InsertCursor
from arcpy.management import Delete, CreateTable, AddFields, CalculateFields
from arcpy.conversion import TableToExcel

from Utils.Helpers import cursor_length

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Alignment


TEMP_TABLE_PATH = "memory/temp_area_diff_table"
MAX_DIFF_CODE_BLOCK = """
from math import sqrt
from functools import partial

def diff_formula(root_coefficient, area_coefficient, area):
    return root_coefficient * sqrt(area) + area_coefficient * area

f1 = partial(diff_formula, .3, .005)
f2 = partial(diff_formula, .8, .002)

def calc_max_diff(area):
    return max(f1(area), f2(area))
    
def is_valid_diff(validity):
    return 'כן' if validity else 'לא'"""


def __format_excel_report__(report_path: str):
    report_file: Workbook = load_workbook(report_path)
    report: Worksheet = report_file.active

    report.title = "הפרשי שטחים"
    report["A1"] = "מס'"
    report.sheet_view.rightToLeft = True

    # Format decimal numbers to maximum two digits after separator
    for row in report.iter_cols(min_row=2, min_col=5, max_col=8):
        for cell in row:
            cell.number_format = "#.0#"

    rtl_align = Alignment(horizontal="right")
    for i in range(1, 10):
        column_letter = get_column_letter(i)
        report.column_dimensions[column_letter].bestFit = True  # Space out columns to best fit
        report[f"{column_letter}1"].alignment = rtl_align # Align header row to be right-to-left

    report_file.save(report_path)


def AreasDifference(parcels_layer, output_path):
    if Exists(TEMP_TABLE_PATH):
        Delete(TEMP_TABLE_PATH)

    parcels = [row for row in
               SearchCursor(parcels_layer,
                            ["ParcelNumber", "BlockNumber", "StatedArea", "SHAPE@AREA", "SubBlockNumber"],
                            "StatedArea IS NOT NULL AND StatedArea > 0 AND Shape__Area > 0",
                            sql_clause=(None, "ORDER BY BlockNumber ASC, SubBlockNumber ASC, ParcelNumber ASC"))]
    num_of_parcels = len(parcels)
    AddMessage(f"Fetched {num_of_parcels:,} parcels.")

    CreateTable("memory", TEMP_TABLE_PATH[7:])
    AddFields(TEMP_TABLE_PATH, [
        ["ParcelNumber", "SHORT", "מס' חלקה"],
        ["BlockNumber", "LONG", "מס' גוש"],
        ["SubBlockNumber", "SHORT", "מס' תת-גוש"],
        ["StatedArea", "DOUBLE", "שטח רשום (מ\"ר)"],
        ["CalculatedArea", "DOUBLE", "שטח מחושב (מ\"ר)"],
        ["AreaDiff", "DOUBLE", "הפרש שטחים (מ\"ר)"],
        ["MaxDiff", "DOUBLE", "הפרש שטחים מקסימלי (מ\"ר)"],
        ["IsValid", "TEXT", "אם ההפרש תקין?", 2]
    ])

    with InsertCursor(TEMP_TABLE_PATH,
                      ["ParcelNumber",
                       "BlockNumber",
                       "StatedArea",
                       "CalculatedArea",
                       "SubBlockNumber"]) as cursor:
        for parcel in parcels:
            cursor.insertRow(parcel)

    AddMessage("Calculating...")
    CalculateFields(TEMP_TABLE_PATH,
                    "PYTHON3",
                    [
                        ["AreaDiff", "abs(!StatedArea! - !CalculatedArea!)"],
                        ["MaxDiff", "calc_max_diff(!StatedArea!)"],
                        ["IsValid", 'is_valid_diff(!AreaDiff! <= !MaxDiff!)']
                    ],
                    MAX_DIFF_CODE_BLOCK)
    AddMessage("Finished")

    invalid_parcels = SearchCursor(TEMP_TABLE_PATH, "OID@", "IsValid = 'לא'")
    num_of_invalid_parcels = cursor_length(invalid_parcels)
    if num_of_invalid_parcels == 0:
        AddMessage("None of the selected parcels' areas differ from the valid range.")
    else:
        AddMessage(f"{num_of_invalid_parcels:,} out of {num_of_parcels:,} parcels have invalid area differences.")

    AddMessage("Exporting reports...")
    TableToExcel(TEMP_TABLE_PATH, output_path, Use_field_alias_as_column_header="ALIAS")
    __format_excel_report__(output_path)
    AddMessage("Finished.")


if __name__ == "__main__":
    param0 = arcpy.GetParameterAsText(0)
    param1 = arcpy.GetParameterAsText(1)

    AreasDifference(param0, param1)
